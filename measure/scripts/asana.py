import requests
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ASANA_TOKEN = 'AFGESCHERMD'
PROJECT_GID = '1206859576165995'

headers = {
    'Authorization': f'Bearer {ASANA_TOKEN}'
}

FASES = [
    'Offerte verzonden',
    'Offerte getekend', 
    'Homepage Design Af',
    'Volledig Design Af',
    'Content & laatste feedback',
    'Afgerond'
]

def get_project_tasks():
    url = f'https://app.asana.com/api/1.0/projects/{PROJECT_GID}/tasks'
    alle_taken = []
    params = {'limit': 100}
    
    while url:
        res = requests.get(url, headers=headers, params=params)
        data = res.json()
        alle_taken.extend(data['data'])
        
        # Volgende pagina ophalen indien aanwezig
        next_page = data.get('next_page')
        if next_page:
            url = next_page['uri']
            params = {}  # params zitten al in de uri
        else:
            url = None
    
    print(f'Totaal {len(alle_taken)} taken opgehaald')
    return alle_taken

def get_task_stories(task_gid):
    url = f'https://app.asana.com/api/1.0/tasks/{task_gid}/stories'
    res = requests.get(url, headers=headers)
    stories = res.json()['data']
    return [s for s in stories if s['resource_subtype'] == 'section_changed']

def extract_section_name(text):
    try:
        to_part = text.split(' to "')[1]
        return to_part.split('"')[0]
    except:
        return None

def main():
    tasks = get_project_tasks()
    rows = []

    for task in tasks:
        stories = get_task_stories(task['gid'])
        if not stories:
            continue

        fase_tijden = {}
        for story in stories:
            section = extract_section_name(story.get('text', ''))
            if section in FASES:
                ts = datetime.fromisoformat(story['created_at'].replace('Z', '+00:00'))
                fase_tijden[section] = ts

        if not fase_tijden:
            continue

        row = {'Project': task['name']}

        for i, fase in enumerate(FASES[:-1]):
            volgende_fase = FASES[i + 1]
            if fase in fase_tijden and volgende_fase in fase_tijden:
                dagen = (fase_tijden[volgende_fase] - fase_tijden[fase]).total_seconds() / 86400
                row[f'{fase} → {volgende_fase}'] = round(dagen, 1)
            else:
                row[f'{fase} → {volgende_fase}'] = ''

        if 'Offerte verzonden' in fase_tijden and 'Afgerond' in fase_tijden:
            totaal = (fase_tijden['Afgerond'] - fase_tijden['Offerte verzonden']).total_seconds() / 86400
            row['Totaal (dagen)'] = round(totaal, 1)
        else:
            row['Totaal (dagen)'] = ''

        rows.append(row)

    # Excel opbouwen
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Nulmeting Doorlooptijd'

    # Headers
    kolommen = ['Project'] + [f'{FASES[i]} → {FASES[i+1]}' for i in range(len(FASES)-1)] + ['Totaal (dagen)']
    
    header_fill = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col_idx, kolom in enumerate(kolommen, 1):
        cel = ws.cell(row=1, column=col_idx, value=kolom)
        cel.fill = header_fill
        cel.font = header_font
        cel.alignment = Alignment(horizontal='center', wrap_text=True)

    # Data
    for row_idx, row in enumerate(rows, 2):
        for col_idx, kolom in enumerate(kolommen, 1):
            waarde = row.get(kolom, '')
            cel = ws.cell(row=row_idx, column=col_idx, value=waarde)
            cel.alignment = Alignment(horizontal='center')
            
            # Afwisselende rijkleur
            if row_idx % 2 == 0:
                cel.fill = PatternFill(start_color='EEF2FF', end_color='EEF2FF', fill_type='solid')

    # Gemiddelden onderaan
    if rows:
        gem_row = len(rows) + 2
        ws.cell(row=gem_row, column=1, value='Gemiddelde').font = Font(bold=True)
        
        for col_idx in range(2, len(kolommen) + 1):
            col_letter = get_column_letter(col_idx)
            ws.cell(row=gem_row, column=col_idx, 
                   value=f'=AVERAGE({col_letter}2:{col_letter}{len(rows)+1})').font = Font(bold=True)

    # Kolombreedtes
    ws.column_dimensions['A'].width = 35
    for col_idx in range(2, len(kolommen) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    ws.row_dimensions[1].height = 40

    wb.save('nulmeting.xlsx')
    print(f'{len(rows)} projecten geëxporteerd naar nulmeting.xlsx')

if __name__ == '__main__':
    main()
